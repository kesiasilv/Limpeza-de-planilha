from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils import get_column_letter

# Caminho para sua planilha original
input_filename = 'Pasta2 (version 1).xlsx'
output_filename = 'planilha_tratada_com_formato_e_links_originais.xlsx' # Novo nome para o arquivo de saída

# --- Passo 1: Carregar a planilha e identificar linhas vazias ---
wb = load_workbook(input_filename)
ws = wb.active

rows_to_keep_indices = []

for r_idx in range(1, ws.max_row + 1):
    row_has_content = False
    max_col_to_check = max(ws.max_column, 20) # Ajuste se sua planilha for mais larga
    
    for c_idx in range(1, max_col_to_check + 1):
        cell = ws.cell(row=r_idx, column=c_idx)
        if cell.value is not None and str(cell.value).strip() != '':
            row_has_content = True
            break

    if row_has_content:
        rows_to_keep_indices.append(r_idx)

# --- Passo 2: Criar uma nova planilha e copiar apenas as linhas desejadas com formatação ---
new_wb = Workbook()
new_ws = new_wb.active
new_ws.title = ws.title

# Copiar larguras de coluna
for col_idx in range(1, ws.max_column + 1):
    letter = get_column_letter(col_idx)
    if letter in ws.column_dimensions:
        new_ws.column_dimensions[letter].width = ws.column_dimensions[letter].width

# Manter as mesclagens (parte mais crítica)
original_merged_cells = []
for merged_range_obj in ws.merged_cells:
    try:
        if isinstance(merged_range_obj, CellRange):
            merged_range_str = str(merged_range_obj)
        elif isinstance(merged_range_obj, str):
            merged_range_str = merged_range_obj
        else:
            continue

        cr = CellRange(merged_range_str)
        original_merged_cells.append(cr)
    except Exception as e:
        print(f"Aviso: Não foi possível processar o intervalo mesclado '{merged_range_obj}'. Pulando. Erro: {e}")
        continue

# Mapeia as linhas originais para as novas linhas na planilha filtrada
row_map = {}
new_row_idx = 1
for old_row_idx in rows_to_keep_indices:
    row_map[old_row_idx] = new_row_idx
    new_row_idx += 1

# Copia os dados e formatação das células relevantes
current_write_row = 1
for original_row_idx in rows_to_keep_indices:
    # Copiar altura da linha para a nova planilha
    if original_row_idx in ws.row_dimensions:
        new_ws.row_dimensions[current_write_row].height = ws.row_dimensions[original_row_idx].height

    # Copiar células e seus estilos
    # Certifique-se de que ws.max_column é grande o suficiente para cobrir todas as colunas com links
    for col_idx in range(1, ws.max_column + 1): 
        original_cell = ws.cell(row=original_row_idx, column=col_idx)
        new_cell = new_ws.cell(row=current_write_row, column=col_idx)

        new_cell.value = original_cell.value

        if original_cell.has_style:
            new_cell.font = original_cell.font.copy()
            new_cell.fill = original_cell.fill.copy()
            new_cell.border = original_cell.border.copy()
            new_cell.alignment = original_cell.alignment.copy()
            new_cell.number_format = original_cell.number_format
        
        # ADIÇÃO CHAVE AQUI: Copiar o hiperlink
        if original_cell.hyperlink is not None:
            # O objeto hyperlink do openpyxl pode ser atribuído diretamente
            new_cell.hyperlink = original_cell.hyperlink
            # Opcional: openpyxl geralmente aplica o estilo de link, mas você pode forçar
            # se não estiver vendo o azul e sublinhado (mas seu problema era o contrário).
            # new_cell.font = Font(color="0000FF", underline="single") 
        
    current_write_row += 1

# Recriar mesclagens
for original_mr in original_merged_cells:
    if original_mr.min_row in row_map and original_mr.max_row in row_map:
        new_min_row = row_map[original_mr.min_row]
        new_max_row = row_map[original_mr.max_row]
        
        new_merged_range = CellRange(
            min_col=original_mr.min_col, max_col=original_mr.max_col,
            min_row=new_min_row, max_row=new_max_row
        )
        
        all_intermediate_rows_kept = True
        for r in range(original_mr.min_row, original_mr.max_row + 1):
            if r not in row_map:
                all_intermediate_rows_kept = False
                break
        
        if all_intermediate_rows_kept:
            try:
                new_ws.merge_cells(str(new_merged_range))
            except Exception as e:
                print(f"Não foi possível recriar mesclagem '{original_mr}' na nova planilha: {e}")
        else:
            print(f"Mesclagem '{original_mr}' (original) não recriada pois linhas internas vazias foram removidas.")
    else:
        print(f"Mesclagem '{original_mr}' (original) não recriada pois linhas de início/fim foram removidas.")


# Salvar a nova planilha
new_wb.save(output_filename)

print(f"Planilha tratada com formatação, links originais e remoção de linhas vazias salva como '{output_filename}'.")
print("Células mescladas que se estendiam por linhas vazias removidas podem não ter sido recriadas.")